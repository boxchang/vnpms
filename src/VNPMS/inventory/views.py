import json
import os

import openpyxl
import xlwt
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, Http404, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from VNPMS.database import database
from bases.utils import get_invform_status_dropdown
from inventory.forms import OfficeInvForm, InvAppliedHistoryForm, OfficeItemForm, SearchForm, AttachmentForm, \
    ItemSearchForm, ItemModelForm, TemplateEditForm
from inventory.models import ItemType, Item, AppliedForm, FormStatus, AppliedItem, Series, Apply_attachment, \
    ItemCategory, Pic_attachment, Setting, ItemFamily, Template_attachment
from users.models import CustomUser, Unit
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from datetime import datetime, timedelta


# 郵件內容
def send_template_email(subject, action, pk, address):
    if address:
        # 電子郵件內容樣板
        form = AppliedForm.objects.get(pk=pk)

        # 查過去請領的資料
        hists = AppliedForm.objects.filter(requester=form.requester, status=FormStatus.objects.get(status_name="已發放"),
                                           category=form.category).order_by('-apply_date')

        key = "{apply_date}{series}".format(apply_date=form.apply_date.replace('-', ''), series=pk)
        files = Apply_attachment.objects.filter(apply=form)
        for file in files:
            file.files.filename = file.files.name[file.files.name.rindex('/') + 1:]
        email_template = render_to_string('inventory/email_template.html', locals())

        email = EmailMessage(
            subject,  # 電子郵件標題
            email_template,  # 電子郵件內容
            settings.EMAIL_HOST_USER,  # 寄件者
            address  # 收件者
        )
        email.fail_silently = False
        email.content_subtype = 'html'
        email.send()
        print("郵件成功寄出")

# 滾序號
def get_series_number(_key, _key_name):
    obj = Series.objects.filter(key=_key)
    if obj:
        _series = obj[0].series + 1
        obj.update(series=_series, desc=_key_name)
    else:
        _series = 1
        Series.objects.create(key=_key, series=1, desc=_key_name)
    return _series


@login_required
def statistic(request):
    item_map = []

    db = database()
    sql = """select category,item_code,spec,cost_center,unitId,unitName, sum(qty)-sum(received_qty) sum_qty from inventory_appliedform a, inventory_applieditem b, users_unit c
             where a.form_no = b.applied_form_id and a.unit_id = c.id
                and a.status_id in (2,3,7)
                group by category,item_code,spec,cost_center,unitId,unitName having sum_qty > 0"""
    rows = db.select_sql_dict(sql)

    item_map = sorted((list(set((dic["item_code"] for dic in rows)))))
    unit_map = (list(set((dic["unitName"] for dic in rows))))

    html_row = "<table border='1' class='table table-bordered table-striped'>"
    # 部門HEADER
    html_row += "<tr><td style='width:40px'></td><td style='width:200px'></td>"
    for unit_data in unit_map:
        html_row += "<td style='text-align:center;width:50px;'>" + unit_data + "</td>"
    html_row += "<td style='text-align:center;width:50px'>加總</td>"
    html_row += "</tr>"

    # 統計數值
    for item_data in item_map:
        item = Item.objects.get(item_code=item_data)

        if item.spec != "自行輸入":
            html_row += "<tr><td>"+item.item_type.category.category_name+"</td><td>"+item.spec+"</td>"
            sum_qty = 0
            for unit_data in unit_map:
                value = ""
                for row in rows:  # 每種料號對應的部門數量
                    if row["spec"]==item.spec and row["unitName"]==unit_data:
                        value = row["sum_qty"]
                        sum_qty += value
                if value:
                    html_row += "<td style='text-align:right;'>"+str(value)+"</td>"
                else:
                    html_row += "<td style='text-align:right;'>0</td>"
            html_row += "<td style='text-align:right;color:blue;font-weight:bold;'>"+str(sum_qty)+"</td>"
            html_row += "</tr>"
        else:  # 自行輸入
            for row in rows:
                if item.item_code == row["item_code"]:
                    html_row += "<tr><td>" + row['category'] + "</td><td>" + row['spec'] + "</td>"
                    sum_qty = 0
                    for unit_data in unit_map:
                        value = ""
                        for row2 in rows:
                            if row["spec"] == row2["spec"] and row["unitName"] == unit_data:
                                value = row["sum_qty"]
                                sum_qty += value

                        if value:
                            html_row += "<td style='text-align:right;'>" + str(value) + "</td>"
                        else:
                            html_row += "<td style='text-align:right;'>0</td>"
                    html_row += "<td style='text-align:right;color:blue;font-weight:bold;'>" + str(sum_qty) + "</td>"
                    html_row += "</tr>"
    html_row += "</table>"

    return render(request, 'inventory/statistic.html', locals())


# 回主頁
def main(request):
    form = OfficeInvForm()

    return render(request, 'inventory/application.html', locals())


@login_required
def import_excel(request):
    if request.method == 'POST':
        try:
            family = ItemFamily.objects.get(family_name="庶務用品")
            excel_file = request.FILES.get('files1')
            if excel_file:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.worksheets[0]
                for iRow in range(2, sheet.max_row+1):
                    if not sheet.cell(row=iRow, column=2).value:
                        break
                    category = sheet.cell(row=iRow, column=2).value
                    category = ItemCategory.objects.get(category_name=category)
                    type = sheet.cell(row=iRow, column=3).value
                    type = ItemType.objects.get(type_name=type)
                    vendor_code = sheet.cell(row=iRow, column=4).value
                    spec = sheet.cell(row=iRow, column=5).value
                    unit = sheet.cell(row=iRow, column=6).value
                    price = sheet.cell(row=iRow, column=8).value

                    if not Item.objects.filter(spec=spec).exists():
                        item = Item()
                        _key = type.category.catogory_code+type.type_code
                        _key_name = type.type_name
                        series = get_series_number(_key, _key_name)
                        item_code = family.family_code + type.category.catogory_code + type.type_code + str(series).zfill(5)
                        item = Item.objects.update_or_create(item_code=item_code, defaults={'item_type': type, 'vendor_code': vendor_code, 'spec': spec, 'unit': unit,
                                                         'price': price, 'create_by': request.user, 'update_by': request.user})
            action = "DONE"
        except Exception as e:
            print(e)

    return render(request, 'inventory/import.html', locals())


@login_required
def apply_list(request):
    action = ""
    if request.method == 'POST':
        action = request.POST.get('action')
        status = request.POST.get('status')
        start_date = request.POST.get('start_date')
        due_date = request.POST.get('due_date')
        category = request.POST.get('category')
        unit = request.POST.get('unit')
        requester = request.POST.get('requester')
        form = InvAppliedHistoryForm(initial={'start_date': start_date, 'due_date': due_date, 'status': status,
                                              'category': category, 'unit': unit, 'requester': requester})
        if unit:
            form.fields["requester"].queryset = CustomUser.objects.filter(unit=unit, is_active=True).all()
    else:
        form = InvAppliedHistoryForm()

    list = get_form_queryset(request)

    if action == "EXPORT":
        return export_form_xls(list)

    return render(request, 'inventory/list.html', locals())


def get_form_queryset(request):
    # ["取消", "已發放", "退單"])
    exclude_list = [4, 5, 6]
    _start_date = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')
    _due_date = datetime.now().strftime('%Y-%m-%d')
    request.session['start_date'] = _start_date
    request.session['due_date'] = _due_date
    list = AppliedForm.objects.all()

    if request.method == 'POST':
        _status = request.POST['status']
        _start_date = str(request.POST['start_date']).replace('/', '-')
        _due_date = str(request.POST['due_date']).replace('/', '-')
        _category = request.POST['category']
        _unit = request.POST['unit']
        _requester = request.POST['requester']

        if _status:
            request.session['status'] = _status
        else:
            if 'status' in request.session:
                del request.session['status']

        if _category:
            request.session['category'] = _category
        else:
            if 'category' in request.session:
                del request.session['category']

        if _unit:
            request.session['unit'] = _unit
        else:
            if 'unit' in request.session:
                del request.session['unit']

        if _requester:
            request.session['requester'] = _requester
        else:
            if 'requester' in request.session:
                del request.session['requester']

        if _start_date and _due_date:
            request.session['start_date'] = _start_date
            request.session['due_date'] = _due_date

    if request.method == 'GET':
        if 'start_date' in request.session:
            _start_date = request.session['start_date']

        if 'due_date' in request.session:
            _due_date = request.session["due_date"]

    if 'status' in request.session:
        _status = request.session["status"]
        list = list.filter(status=_status)
    else:
        list = list.exclude(status__in=exclude_list)

    if 'category' in request.session:
        _category = request.session["category"]
        list = list.filter(category=_category)

    # 沒有輸入人名條件才會觸發部門條件
    if 'requester' in request.session:
        _requester = request.session["requester"]
        list = list.filter(Q(requester=_requester) | Q(approver=_requester) | Q(create_by=_requester))
    else:
        if 'unit' in request.session:
            _unit = request.session["unit"]
            list = list.filter(unit=_unit)

    if not request.user.has_perm("perm_misc_apply"):  # 不是管理者只能看自己的單據
        list = list.filter(Q(requester=request.user) | Q(approver=request.user) | Q(create_by=request.user))

    list = list.filter(apply_date__gte=_start_date, apply_date__lte=_due_date)
    list = list.order_by('-apply_date', '-form_no')
    return list


@login_required
def approve(request):
    list = AppliedForm.objects.filter(status_id=1, approver=request.user)
    return render(request, 'inventory/approve.html', locals())


def unlock(key):
    apply_date = key[0:4] + "-" + key[4:6] + "-" + key[6:8]
    pk = key[8:]
    return apply_date, pk

@login_required
def agree(request, key):
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        apply = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        apply.status = FormStatus.objects.get(pk=2)
        apply.approve_time = datetime.now()
        apply.save()
    return redirect(reverse('inv_approve'))


def mail_agree(request, key):
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        form = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        items = form.applied_form_item.all().order_by('category')

        # 取消
        if form.status.id == 4:
            action = "cancel"
        # 處理中/退單
        elif form.status.id in [2, 6]:
            action = "done"
        else:
            form.status = FormStatus.objects.get(pk=2)
            form.approve_time = datetime.now()
            form.save()
            action = "agree"

        return render(request, 'inventory/email_template.html', locals())


@login_required
def reject(request, key):
    address = []
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        apply = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        apply.status = FormStatus.objects.get(pk=6)
        apply.approve_time = datetime.now()
        apply.save()

        if apply.requester.email:
            address.append(apply.requester.email)

        attr = Setting.objects.get(attr="reject_mail")
        if attr:
            emails = attr.values.split(';')
            for email in emails:
                address.append(email)

        subject = '總務用品請領單，退單通知!!!!'
        try:
            send_template_email(subject, action="reject", pk=apply.pk, address=address)
        except Exception as e:
            print(e)

    return redirect(reverse('inv_approve'))


def mail_reject(request, key):
    address = []
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        form = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        items = form.applied_form_item.all().order_by('category')

        # 取消
        if form.status.id == 4:
            action = "cancel"
        # 處理中/退單
        elif form.status.id in [2, 6]:
            action = "done"
        else:
            form.status = FormStatus.objects.get(pk=6)
            form.approve_time = datetime.now()
            form.save()

            action = "reject"
            if form.requester.email:
                address.append(form.requester.email)

            if form.create_by.email:
                address.append(form.create_by.email)

            attr = Setting.objects.get(attr="reject_mail")
            if attr:
                emails = attr.values.split(';')
                for email in emails:
                    address.append(email)

            subject = '總務用品請領單，退單通知!!!!'
            try:
                send_template_email(subject, action, pk=form.pk, address=address)
            except Exception as e:
                print(e)

        return render(request, 'inventory/email_template.html', locals())


@login_required
def delete(request, pk):
    if request.method == 'GET':
        apply = AppliedForm.objects.get(pk=pk)
        apply.status = FormStatus.objects.get(pk=4)
        apply.update_by = request.user
        apply.save()

    return redirect(reverse('inv_list'))


@login_required
def apply(request):
    address = []
    if request.method == 'POST':
        hidCart_list = request.POST.get('hidCart_list')
        if hidCart_list:
            items = json.loads(hidCart_list)

        unit = request.POST.get('unit')
        requester = request.POST.get('requester')
        apply_date = request.POST.get('apply_date')
        ext_number = request.POST.get('ext_number')
        reason = request.POST.get('reason')

        try:
            apply = AppliedForm()
            YYYYMM = datetime.now().strftime("%Y%m")
            key = "OR"+YYYYMM
            apply.form_no = key + str(get_series_number(key, "文具請領")).zfill(3)
            apply.unit = Unit.objects.get(pk=unit)
            apply.requester = CustomUser.objects.get(id=requester)
            apply.apply_date = apply_date
            apply.ext_number = ext_number
            apply.reason = reason
            apply.status = FormStatus.objects.get(pk=1)
            apply.category = ItemCategory.objects.filter(category_name=items[0]['category']).first()
            apply.create_by = request.user
            apply.update_by = request.user
            apply.approver = request.user.manager
            apply.save()

            if request.FILES.get('file1'):
                request_file = Apply_attachment(files=request.FILES['file1'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file2'):
                request_file = Apply_attachment(files=request.FILES['file2'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file3'):
                request_file = Apply_attachment(files=request.FILES['file3'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()

            for item in items:
                obj = AppliedItem()
                obj.category = item['category']
                obj.item_code = item['item_code']
                obj.spec = item['spec']
                obj.qty = item['qty']
                obj.unit = item['unit']
                obj.comment = item['comment']
                obj.applied_form = apply
                obj.save()

            if apply.requester.manager.email:
                address.append(apply.requester.manager.email)

            subject = '總務用品請領單簽核通知'
            try:
                send_template_email(subject, action="email", pk=apply.pk, address=address)
            except Exception as e:
                print(e)

        except Exception as e:
            print(e)

        return redirect(apply.get_absolute_url())

    templates = Template_attachment.objects.all()
    form = OfficeInvForm(initial={"unit": request.user.unit, "requester": request.user})
    form.fields["requester"].queryset = CustomUser.objects.filter(unit=request.user.unit, is_active=True).all()
    search_form = SearchForm()
    attach_form = AttachmentForm()
    return render(request, 'inventory/application.html', locals())


@login_required
def detail(request, pk):
    try:
        form = AppliedForm.objects.get(pk=pk)
        items = form.applied_form_item.all().order_by('category')

        status_html = get_invform_status_dropdown(form)

        if form.status.id == 1 and form.requester.manager == request.user:
            isApprover = True

        # 只有簽核中的狀態才能取消
        if form.status.id == 1 and form.create_by == request.user:
            isCreater = True

        files = Apply_attachment.objects.filter(apply=form)
        for file in files:
            file.files.filename = file.files.name[file.files.name.rindex('/')+1:]
        key = "{apply_date}{series}".format(apply_date=form.apply_date.replace('-', ''),
                                           series=form.pk)
        for form_item in items:
            form_item.x = Item.objects.get(item_code=form_item.item_code)

        # 查過去請領的資料
        two_year_ago = datetime.now() - timedelta(days=730)
        hists = AppliedForm.objects.filter(requester=form.requester, create_at__gte=two_year_ago,
                                           status=FormStatus.objects.get(status_name="已發放"), category=form.category).order_by('-apply_date')

    except AppliedForm.DoesNotExist:
        raise Http404('Form does not exist')


    return render(request, 'inventory/detail.html', locals())


#API類別
def TypeAPI(request, category_id):
    type_data = ItemType.objects.filter(category_id=int(category_id)).values('id', 'type_name').order_by('type_code')
    type_list = []
    for data in type_data:
        type_list.append({'id': data['id'], 'type_name': data['type_name']})

    return JsonResponse(type_list, safe = False)


#API類別
def ItemAPI(request):
    item_list = []
    if request.method == 'POST':
        category_id = request.POST.get('category_id')
        type_id = request.POST.get('type_id')
        keyword = request.POST.get('keyword')

        item_data = Item.objects.filter(enabled=True)

        if type_id:
            item_data = item_data.values('item_code', 'spec', 'unit', 'item_pics__files')
        else:
            item_data = item_data.exclude(spec__contains="自行輸入").values('item_code', 'spec', 'unit', 'item_pics__files')

        if category_id:
            item_type = ItemType.objects.filter(category_id=category_id)
            item_data = item_data.filter(item_type__in=item_type).values('item_code', 'spec', 'unit', 'item_pics__files')
        if type_id:
            item_data = item_data.filter(item_type_id=int(type_id)).values('item_code', 'spec', 'unit', 'item_pics__files')
        if keyword:
            query = Q(spec__icontains=keyword)
            item_data = item_data.filter(query).values('item_code', 'spec', 'price', 'unit', 'item_pics__files')

        for data in item_data:
            category = ItemCategory.objects.get(id=category_id)
            item_list.append({'item_code': data['item_code'], 'spec': data['spec'], 'unit': data['unit'],
                              'category': category.category_name, 'pic': data['item_pics__files']})

    return JsonResponse(item_list, safe=False)


def change_status(request):
    if request.POST:
        form_id = request.POST.get('form_id')
        status_id = request.POST.get('status_id')

        status = FormStatus.objects.get(pk=status_id)
        obj = AppliedForm.objects.get(pk=form_id)
        obj.status = status
        obj.save()

        return redirect(obj.get_absolute_url())


def recieved(request, pk):
    item = AppliedItem.objects.get(pk=pk)
    if request.POST:
        applied_form = AppliedForm.objects.get(pk=item.applied_form.pk)
        form = OfficeItemForm(request.POST, instance=item)
        form.save()
        return redirect(applied_form.get_absolute_url())

    form = OfficeItemForm(instance=item)
    return render(request, 'inventory/recieved.html', locals())


def pr_apply(request):
    if request.method == 'POST':
        hidCart_list = request.POST.get('hidCart_list')
        if hidCart_list:
            items = json.loads(hidCart_list)

        unit = request.POST.get('unit')
        requester = request.POST.get('requester')
        apply_date = request.POST.get('apply_date')
        ext_number = request.POST.get('ext_number')
        reason = request.POST.get('reason')

        try:
            apply = AppliedForm()
            apply.unit = Unit.objects.get(pk=unit)
            apply.requester = CustomUser.objects.get(id=requester)
            apply.apply_date = apply_date
            apply.ext_number = ext_number
            apply.reason = reason
            apply.status = FormStatus.objects.get(pk=1)
            apply.create_by = request.user
            apply.save()

            if request.FILES.get('file1'):
                request_file = Apply_attachment(files=request.FILES['file1'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file2'):
                request_file = Apply_attachment(files=request.FILES['file2'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file3'):
                request_file = Apply_attachment(files=request.FILES['file3'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()

            for item in items:
                obj = AppliedItem()
                obj.item_code = item['item_code']
                obj.spec = item['spec']
                obj.qty = item['qty']
                obj.unit = item['unit']
                obj.applied_form = apply
                obj.save()

            action = "email"
            email = apply.requester.manager.email
            if email:
                # 電子郵件內容樣板
                pk = apply.pk
                form = AppliedForm.objects.get(pk=pk)
                files = Apply_attachment.objects.filter(apply=form)
                for file in files:
                    file.files.filename = file.files.name[file.files.name.rindex('/') + 1:]
                email_template = render_to_string('inventory/email_template.html', locals())

                email = EmailMessage(
                    '總務用品請領單簽核通知',  # 電子郵件標題
                    email_template,  # 電子郵件內容
                    settings.EMAIL_HOST_USER,  # 寄件者
                    [email]  # 收件者
                )
                email.fail_silently = False
                email.content_subtype = 'html'
                email.send()
                print("郵件成功寄出")

        except Exception as e:
            print(e)

        return redirect(apply.get_absolute_url())

    form = OfficeInvForm()
    search_form = SearchForm()
    attach_form = AttachmentForm()
    return render(request, 'inventory/pr_apply.html', locals())


def search(request):
    search_form = SearchForm()
    return render(request, 'inventory/search.html', locals())


@login_required
def item_list(request):
    page_number = 1
    items = Item.objects.all()
    search_form = ItemSearchForm()
    pic = ""
    category_id = ""
    type_id = ""
    keyword = ""

    if request.method == 'POST':
        category_id = request.POST.get('category')
        type_id = request.POST.get('type')
        keyword = request.POST.get('keyword')
        pic = request.POST.get('pic')

    if request.method == 'GET':
        if request.GET.get('page'):
            page_number = int(request.GET.get('page'))

        if 'pic' in request.session:
            pic = request.session['pic']

        if 'category_id' in request.session:
            category_id = request.session['category_id']

        if 'type_id' in request.session:
            type_id = request.session['type_id']

        if 'keyword' in request.session:
            keyword = request.session['keyword']

    if pic == "True":
        request.session['pic'] = pic
        items = items.filter(item_pics__isnull=False)
    elif pic == "False":
        request.session['pic'] = pic
        items = items.filter(item_pics__isnull=True)
    else:
        if 'pic' in request.session:
            del request.session['pic']

    if category_id:
        request.session['category_id'] = category_id
        item_type = ItemType.objects.filter(category_id=category_id)
        items = items.filter(item_type__in=item_type)
    else:
        if 'category_id' in request.session:
            del request.session['category_id']

    if type_id:
        request.session['type_id'] = type_id
        items = items.filter(item_type_id=int(type_id))
    else:
        if 'type_id' in request.session:
            del request.session['type_id']

    if keyword:
        request.session['keyword'] = keyword
        items = items.filter(Q(item_code__contains=keyword) | Q(spec__icontains=keyword))
    else:
        if 'keyword' in request.session:
            del request.session['keyword']

    search_form = ItemSearchForm(initial={'type': type_id, 'category': category_id, 'keyword': keyword})
    if category_id:
        search_form.fields['type'].queryset = ItemType.objects.filter(category_id=category_id)

    results = list(items)
    page_obj = Paginator(results, 15)

    if page_number:
        page_results = page_obj.page(page_number)
    else:
        page_results = page_obj.page(1)

    return render(request, 'inventory/item_list.html', locals())


@login_required
def item_detail(request, pk):
    try:
        item = Item.objects.get(pk=pk)
        pics = Pic_attachment.objects.filter(item=item).all()
    except Item.DoesNotExist:
        raise Http404('Item does not exist')

    return render(request, 'inventory/item_detail.html', locals())


@login_required
def item_update(request, pk):
    mode = "UPDATE"
    item = Item.objects.get(id=pk)
    if request.method == "POST":
        form = ItemModelForm(request.POST, instance=item)
        if form.is_valid():
            tmp_form = form.save(commit=False)
            tmp_form.update_by = request.user
            tmp_form.save()

            if request.FILES.get('pic1'):
                Pic_attachment.objects.update_or_create(item=tmp_form, defaults={'files': request.FILES['pic1'], 'create_by': request.user})
            return redirect(tmp_form.get_absolute_url())
    else:
        form = ItemModelForm(instance=item)

    return render(request, 'inventory/item_edit.html', locals())


@login_required
def item_create(request):
    if request.method == "POST":
        form = ItemModelForm(request.POST)
        if form.is_valid():
            tmp_form = form.save(commit=False)
            _key = tmp_form.item_type.category.family.family_code + tmp_form.item_type.category.catogory_code + tmp_form.item_type.type_code
            _key_name = tmp_form.item_type.type_name
            series = get_series_number(_key, _key_name)
            item_code = _key + str(series).zfill(5)
            tmp_form.item_code = item_code
            tmp_form.create_by = request.user
            tmp_form.update_by = request.user
            form.save()

            if request.FILES.get('pic1'):
                request_file = Pic_attachment(
                    files=request.FILES['pic1'])
                request_file.item = tmp_form
                request_file.create_by = request.user
                request_file.save()

            return redirect(tmp_form.get_absolute_url())
    else:
        form = ItemModelForm()

    return render(request, 'inventory/item_edit.html', locals())


@login_required
def setting(request):
    if request.method == 'POST':
        reject_mail = request.POST.get('reject_mail')
        attrs = Setting.objects.update_or_create(attr="reject_mail", defaults={"values": reject_mail})
        result = "DONE"
        print(reject_mail)

    try:
        attr = Setting.objects.get(attr="reject_mail")
        if attr:
            reject_mail = attr.values
    except Exception as e:
        print(e)

    return render(request, 'inventory/setting.html', locals())

#Excel
def export_form_xls(list):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="inventory_application.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Forms')
    ws.col(0).width = 256 *20
    ws.col(1).width = 256 *20
    ws.col(2).width = 256 *20
    ws.col(3).width = 256 *20
    ws.col(4).width = 256 *20
    ws.col(5).width = 256 *20
    ws.col(6).width = 256 *20

    # Sheet header, first row
    row_num = 0
    form_columns = ['單號', '申請日期', '狀態', '簽核者', '申請部門', '申請者', '分機']
    form_item_columns = ['', '物品類別', '品名', '申請數量', '已發放數量', '單位', '備註']

    # Title Style
    title_style = xlwt.XFStyle()
    title_style.font.bold = True
    title_style.font.colour_index = xlwt.Style.colour_map['blue']
    title_style.font.height = 20 * 14

    # Form Header Style
    header_style = xlwt.XFStyle()
    header_style.font.bold = True
    header_style.font.colour_index = xlwt.Style.colour_map['white']
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = xlwt.Style.colour_map['gray80']
    header_style.pattern = pattern

    # Header Content Style
    hc_style = xlwt.XFStyle()
    hc_style.font.colour_index = xlwt.Style.colour_map['white']
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = xlwt.Style.colour_map['gray50']
    hc_style.pattern = pattern

    # Form Item Header
    item_header_style = xlwt.XFStyle()
    item_header_style.font.bold = True
    item_header_style.font.colour_index = xlwt.Style.colour_map['white']
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = xlwt.Style.colour_map['gray40']
    item_header_style.pattern = pattern

    # Normal Style
    font_style = xlwt.XFStyle()
    font_style.font.colour_index = xlwt.Style.colour_map['black']

    # Title
    ws.write(row_num, 0, "台灣伊格爾博格曼股份有限公司", title_style)
    row_num += 1
    ws.write(row_num, 0, "總務用品請領單", title_style)
    row_num += 1

    for data in list:
        approver = ""
        if data.approver:
            approver = data.approver.username

        for col_num in range(len(form_columns)):
            ws.write(row_num, col_num, form_columns[col_num], header_style)

        row_num += 1
        # Sheet body, remaining rows
        ws.write(row_num, 0, data.form_no, hc_style)
        ws.write(row_num, 1, data.apply_date, hc_style)
        ws.write(row_num, 2, data.status.status_name, hc_style)
        ws.write(row_num, 3, approver, hc_style)
        ws.write(row_num, 4, data.unit.unitName, hc_style)
        ws.write(row_num, 5, data.requester.username, hc_style)
        ws.write(row_num, 6, data.ext_number, hc_style)

        # 申請原因
        row_num += 1
        ws.write(row_num, 0, "申請原因：", font_style)
        ws.write(row_num, 1, data.reason, font_style)

        items = data.applied_form_item.all().order_by('category')

        # Form Item Header
        row_num += 1
        for col_num in range(len(form_item_columns)):
            ws.write(row_num, col_num, form_item_columns[col_num], item_header_style)

        for form_item in items:
            row_num += 1
            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()
            ws.write(row_num, 0, "", font_style)
            ws.write(row_num, 1, form_item.category, font_style)
            ws.write(row_num, 2, form_item.spec, font_style)
            ws.write(row_num, 3, form_item.qty, font_style)
            ws.write(row_num, 4, form_item.received_qty, font_style)
            ws.write(row_num, 5, form_item.unit, font_style)
            ws.write(row_num, 6, form_item.comment, font_style)
        row_num += 2

    wb.save(response)
    return response

def CategoryAPI(request, family_id):
    category_data = ItemCategory.objects.filter(family_id=int(family_id)).values('id', 'category_name')
    category_list = []
    for data in category_data:
        category_list.append({'id': data['id'], 'category_name': data['category_name']})
    return JsonResponse(category_list, safe=False)


def TypeAPI(request, category_id):
    type_data = ItemType.objects.filter(category_id=int(category_id)).values('id', 'type_name')
    type_list = []
    for data in type_data:
        type_list.append({'id': data['id'], 'type_name': data['type_name']})
    return JsonResponse(type_list, safe=False)


def template_edit(request):
    result = ""
    if request.method == 'POST':
        attach_form = TemplateEditForm(request.POST, request.FILES)
        if attach_form.is_valid():
            files = Template_attachment.objects.all()

            key_file = request.FILES['key_file'] if 'key_file' in request.FILES else None
            stamp_file = request.FILES['stamp_file'] if 'stamp_file' in request.FILES else None
            print_file = request.FILES['print_file'] if 'print_file' in request.FILES else None

            if files:
                if key_file:
                    os.remove(files[0].key_file.path)
                    files[0].key_file = key_file

                if stamp_file:
                    os.remove(files[0].stamp_file.path)
                    files[0].stamp_file = stamp_file

                if print_file:
                    os.remove(files[0].print_file.path)
                    files[0].print_file = print_file

                files[0].save()
            else:
                Template_attachment.objects.create(key_file=key_file, stamp_file=stamp_file, print_file=print_file, update_by=request.user)
            result = "DONE"
    attach_form = TemplateEditForm()
    return render(request, 'inventory/template_edit.html', locals())