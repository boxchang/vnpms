import csv
import json
import os
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.db.models import Sum, Max
from django.http import JsonResponse
from django.shortcuts import render, redirect
import openpyxl
import uuid
from datetime import datetime
import xlwt
from django.urls import reverse
from django.utils.translation import get_language
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from VNPMS.database import database, dc_database
from VNPMS.settings.base import MEDIA_ROOT
from bases.utils import get_date_str
from jobs.sap_sync.SYN_Noah_Consumption import SYN_Noah_Consumption
from jobs.sap_sync.SYN_Noah_WorkHour import SYN_Noah_WorkHour
from production.encode import get_series_number
from production.forms import RecordForm, RecordSearchForm, WoSearchForm, RecordManageForm, ExportForm, \
    RecordHistoryForm, ItemSearchForm
from production.models import ExcelTemp, WODetail, Record, Record2, WorkType, COOIS_Record, WOMain, Machine, \
    Consumption, Sync_SAP_Log
from users.models import CustomUser
from django.utils.translation import gettext_lazy as _
from django.conf import settings

def COOIS2Table(items):
    html = """<table border='1' class='table table-bordered table-striped'>
                    {Rows}
                </table>"""
    sRow = ""
    sCol = ""
    sCol += """<th>序號</th><th>工廠</th><th>訂單</th><th>確認</th><th>作業</th><th>Text key</th><th>作業短文</th><th>控制碼</th><th>系統狀態</th>
               <th>工作中心</th><th>料號</th><th>物料說明</th><th>作業數量</th><th>人時</th><th>機時</th><th>基礎數量</th>"""
    sCol = "<tr>" + sCol + "</tr>"
    sRow += sCol
    index = 1
    for item in items:
        sCol = ""
        sCol += "<td>{value}</td>".format(value=index)
        sCol += "<td>{value}</td>".format(value=item['plant'])
        sCol += "<td>{value}</td>".format(value=item['wo_no'])
        sCol += "<td>{value}</td>".format(value=item['cfm_code'])
        sCol += "<td>{value}</td>".format(value=item['step_no'])
        sCol += "<td>{value}</td>".format(value=item['step_code'])
        sCol += "<td>{value}</td>".format(value=item['step_name'])
        sCol += "<td>{value}</td>".format(value=item['ctr_code'])
        sCol += "<td>{value}</td>".format(value=item['status'])
        sCol += "<td>{value}</td>".format(value=item['work_center'])
        sCol += "<td>{value}</td>".format(value=item['item_no'])
        sCol += "<td>{value}</td>".format(value=item['spec'])
        sCol += "<td>{value}</td>".format(value=item['wo_qty'])
        sCol += "<td>{value}</td>".format(value=item['wo_labor_time'])
        sCol += "<td>{value}</td>".format(value=item['wo_mach_time'])
        sCol += "<td>{value}</td>".format(value=item['std_qty'])
        sCol = "<tr>" + sCol + "</tr>"
        sRow += sCol

        index += 1
    html = html.format(Rows=sRow)
    return html


# 新增報工
def record(request):
    error_msg = ""
    if request.method == 'POST':
        mtr_info = request.POST.get('hid_mtr_info')
        record_dt = request.POST.get('record_dt')
        emp_no = request.POST.get('emp_no')
        username = request.POST.get('username')
        sap_emp_no = request.POST.get('sap_emp_no')
        wo_no = request.POST.get('wo_no')
        work_center = request.POST.get('work_center')
        item_no = request.POST.get('item_no')
        spec = request.POST.get('spec')
        cfm_code = request.POST.get('cfm_code')
        ctr_code = request.POST.get('ctr_code')
        labor_time = request.POST.get('labor_time')
        mach_time = request.POST.get('mach_time')
        good_qty = request.POST.get('good_qty')
        ng_qty = request.POST.get('ng_qty')
        step_no = request.POST.get('step_no')
        step_code = request.POST.get('step_code')
        step_name = request.POST.get('step_name')
        plant = request.POST.get('plant')
        comment = request.POST.get('comment')
        request.session['record_dt'] = record_dt
        key_user = CustomUser.objects.get(sap_emp_no=sap_emp_no)
        mach_code = request.POST.get('mach_code')
        status = request.POST.get('status')
        mach = None
        key = get_date_str()
        # record = Record.objects.update_or_create(record_dt=record_dt, emp_no=emp_no, wo_no=wo_no, cfm_code=cfm_code,
        #                                 defaults={'labor_time': labor_time, 'mach_time': mach_time, 'ctr_code': ctr_code,
        #                                           'good_qty': good_qty, 'ng_qty': ng_qty, 'spec':spec, 'username': username,
        #                                           'step_code': step_code, 'step_name': step_name, 'sap_emp_no': sap_emp_no,
        #                                           'update_by': user})

        if mach_code:
            mach = Machine.objects.get(mach_code=mach_code)
        _series = get_series_number("record", key)
        series_no = "T" + key + str(_series).zfill(5)
        record = Record.objects.create(record_dt=record_dt, emp_no=emp_no, wo_no=wo_no, cfm_code=cfm_code,
                                        labor_time=labor_time, mach_time=mach_time, ctr_code=ctr_code,
                                        good_qty=good_qty, ng_qty=ng_qty, item_no=item_no, spec=spec, username=username,
                                        step_no=step_no, step_code=step_code, step_name=step_name, sap_emp_no=sap_emp_no,
                                        update_by=key_user, plant=plant, work_center=work_center, comment=comment,
                                        status=status, mach=mach, id=series_no)
        
        # 302B 物料耗用資料
        if mtr_info:
            mtr_info = json.loads(mtr_info)
            for mtr in mtr_info:
                _series = get_series_number("consumption", key)
                series_no = "M" + key + str(_series).zfill(5)
                Consumption.objects.create(plant=plant, cfm_code=cfm_code, wo_no=wo_no, item_no=mtr['mtr_no'], qty=mtr['qty'], create_by=key_user, id=series_no, wo_mtrl_no=mtr['wo_mtrl_no'])

        return redirect(record.get_absolute_url())

    form = RecordForm()
    return render(request, 'production/record.html', locals())


@csrf_exempt
def record2(request):
    if request.method == 'POST':
        sap_emp_no = request.POST.get('hid_sap_emp_no')
        record_dt = request.POST.get('hid_record_dt2')
        work_type = request.POST.get('work_type')
        qty = request.POST.get('qty')
        key = get_date_str()

        key_user = CustomUser.objects.get(sap_emp_no=sap_emp_no)
        if work_type:
            work_type = WorkType.objects.get(type_code=work_type)

        if not qty:
            qty = 0

        comment = request.POST.get('comment')
        labor_time = request.POST.get('labor_time')
        # record2 = Record2.objects.update_or_create(record_dt=record_dt, sap_emp_no=sap_emp_no, work_type=work_type,
        #                                          defaults={'labor_time': labor_time,
        #                                                    'comment': comment,
        #                                                    'create_by': key_user})
        _series = get_series_number("record2", key)
        series_no = "T" + key + str(_series).zfill(5)
        record2 = Record2.objects.create(record_dt=record_dt, sap_emp_no=sap_emp_no, work_type=work_type,
                                         labor_time=labor_time, comment=comment, create_by=key_user, id=series_no, qty=qty)
        return redirect(record2.get_absolute_url())
    return render(request, 'production/record.html', locals())


# 工單報工資料查詢
def wo_detail(request):
    if request.method == 'POST':
        wo_no = request.POST.get('wo_no')
        if wo_no:
            wo_no = str(wo_no).strip()
        steps = WODetail.objects.select_related('wo_main').filter(wo_main__wo_no=wo_no, wo_main__enable=True).order_by('step_no')
        for step in steps:
            step_labor_time = 0
            step_mach_time = 0
            step_good_qty = 0
            step_ng_qty = 0
            records = Record.objects.filter(wo_no=step.wo_main.wo_no, step_no=step.step_no)
            for record in records:
                step_labor_time += record.labor_time
                step_mach_time += record.mach_time
                step_good_qty += record.good_qty
                step_ng_qty += record.ng_qty
            step.work_records = records
            step.step_labor_time = step_labor_time
            step.step_mach_time = step_mach_time
            step.step_good_qty = step_good_qty
            step.step_ng_qty = step_ng_qty
            step.std_labor_time = round(step.wo_labor_time/step.std_qty*step.wo_qty, 1)
            step.std_mach_time = round(step.wo_mach_time / step.std_qty * step.wo_qty, 1)
            item_no = step.wo_main.item_no
            spec = step.wo_main.spec

    form = WoSearchForm()
    return render(request, 'production/wo_detail.html', locals())


# 料號查詢工單報工資料
def item_search(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        due_date = request.POST.get('due_date')
        item_no = request.POST.get('item_no')
        if item_no:
            item_no = str(item_no).strip()
        wos = WOMain.objects.filter(item_no=item_no, enable=True, create_at__gte=start_date, create_at__lte=due_date).order_by('-create_at')

    form = ItemSearchForm()
    return render(request, 'production/item_search.html', locals())


# 報工刪除
@login_required
def record_del(request, pk):
    record = Record.objects.get(pk=pk)
    consumption = Consumption.objects.filter(cfm_code=record.cfm_code)
    consumption.delete()
    record.delete()
    return redirect(reverse('prod_record_detail_sap_empno', kwargs={'sap_emp_no': record.sap_emp_no}))


# 其他報工刪除
@login_required
def record2_del(request, pk):
    record = Record2.objects.get(pk=pk)
    record.delete()
    return redirect(reverse('prod_record_detail_sap_empno', kwargs={'sap_emp_no': record.sap_emp_no}))


# 報工資料編輯
def record_edit(request, pk):
    record = Record.objects.filter(pk=pk)
    record_dt = record[0].record_dt
    sap_emp_no = record[0].sap_emp_no
    step_code = record[0].step_code
    mach_code = record[0].mach
    if request.method == 'POST':
        labor_time = request.POST.get('labor_time')
        mach_time = request.POST.get('mach_time')
        good_qty = request.POST.get('good_qty')
        ng_qty = request.POST.get('ng_qty')
        emp_no = request.POST.get('emp_no')
        username = request.POST.get('username')
        wo_no = request.POST.get('wo_no')
        work_center = request.POST.get('work_center')
        item_no = request.POST.get('item_no')
        spec = request.POST.get('spec')
        cfm_code = request.POST.get('cfm_code')
        ctr_code = request.POST.get('ctr_code')
        step_no = request.POST.get('step_no')
        step_code = request.POST.get('step_code')
        step_name = request.POST.get('step_name')
        comment = request.POST.get('comment')
        plant = request.POST.get('plant')
        mach_code = request.POST.get('mach_code')
        mach = None
        if mach_code:
            mach = Machine.objects.get(mach_code=mach_code)

        record.update(labor_time=labor_time, mach_time=mach_time, good_qty=good_qty, ng_qty=ng_qty, item_no=item_no,
                                            spec=spec, username=username, wo_no=wo_no,
                                            step_no=step_no, step_code=step_code, step_name=step_name, sap_emp_no=sap_emp_no,
                                            update_by=request.user, plant=plant, work_center=work_center, comment=comment, mach=mach)

        form = RecordSearchForm(initial={'record_dt': record_dt})

        return redirect(reverse('prod_record_detail_sap_empno', kwargs={'sap_emp_no': sap_emp_no}))

    # 已累計報工
    worked_labor_time = 0
    records = Record.objects.filter(record_dt=record_dt, sap_emp_no=sap_emp_no).aggregate(Sum('labor_time'),
                                                                                  Sum('mach_time'))
    if records['labor_time__sum']:
        worked_labor_time = records['labor_time__sum']


    form = RecordForm(instance=record[0], initial={'worked_labor_time': worked_labor_time, 'mach_code': mach_code})
    form.fields["mach_code"].queryset = Machine.objects.filter(step_code=step_code).all()

    return render(request, 'production/record.html', locals())


# 報工資料查詢
def record_detail(request):
    sap_emp_no = ""
    if request.method == 'POST':
        sap_emp_no = request.POST['sap_emp_no']

    return record_detail_sap_empno(request, sap_emp_no)


# 報工資料查詢
def record_detail_sap_empno(request, sap_emp_no):
    html = ""
    username = ""
    worktypes = WorkType.objects.all().order_by('type_code')
    lang = get_language()
    if request.method == 'POST':
        sap_emp_no = request.POST['sap_emp_no']
        record_dt = request.POST['record_dt']
        request.session['record_dt'] = record_dt
    else:
        if 'record_dt' in request.session:
            record_dt = request.session['record_dt']
        else:
            now = datetime.now()
            record_dt = datetime.strftime(now, '%Y-%m-%d')

    if sap_emp_no:
        try:
            key_user = CustomUser.objects.get(sap_emp_no=sap_emp_no)
        except CustomUser.DoesNotExist:
            key_user = None

        if key_user:
            username = key_user.username
        total_labor_time = 0
        rest_time = 0
        records = Record.objects.filter(record_dt=record_dt, sap_emp_no=sap_emp_no)
        record2s = Record2.objects.filter(record_dt=record_dt, sap_emp_no=sap_emp_no)
        table = "<table border='1' class='table table-bordered'>{header}{body}</table>"
        header = """<tr style='background-color:#EEE;'>
                        <th style='width:100px'>{prod_order}</th>
                        <th style='width:100px'>{item_no}</th>
                        <th style='width:200px'>{spec}</th>
                        <th style='width:100px'>{step_code}</th>
                        <th style='width:100px'>{step_name}</th>
                        <th style='text-align:center;width:100px'>{labor_time}</th>
                        <th style='text-align:center;width:100px'>{mach_time}</th>
                        <th style='text-align:center;width:100px'>{good_qty}</th>
                        <th style='text-align:center;width:100px'>NG</th>
                        <th style='width:150px'></th></tr>""".format(
            prod_order=_('prod_order'), step_code=_('step_code'), step_name=_('step_name'),
            labor_time=_('labor_time'), mach_time=_('mach_time'), good_qty=_('good_qty'),
            item_no=_('item_no'), spec=_('spec'))
        body = ""
        for record in records:
            total_labor_time += record.labor_time
            body_tmp = """<tr>
                            <td rowspan='2' height='120px'>{wo_no}</td>
                            <td rowspan='2'>{item_no}</td>
                            <td rowspan='2'>{spec}</td>
                            <td height='60px'>{step_code}</td><td>{step_name}</td>
                            <td style='text-align:right'>{labor_time}</td>
                            <td style='text-align:right'>{mach_time}{mach_name}</td>
                            <td style='text-align:right'>{good_qty}</td>
                            <td style='text-align:right'>{ng_qty}</td><td>{edit_btn}{del_btn}</td></tr>"""
            body_tmp += """<tr>
                            <td colspan='5'>{comment}</td>
                            <td></td>
                           </tr>"""
            if not request.user.is_anonymous:
                edit_btn = """<a class=\"btn btn-info m-1\" href=\"/production\\record_edit\\{record_pk}\" role=\"button\">編輯</a>""".format(
                    record_pk=record.pk)
                del_btn = """<a class=\"btn btn-danger m-1\" href=\"/production\\record_delete\\{record_pk}\" role=\"button\" onclick=\"return confirm('Are you sure?')\">刪除</a>""".format(
                    record_pk=record.pk)
            else:
                edit_btn = ""
                del_btn = ""
            comment = "" if record.comment is None else record.comment
            mach_name = "" if record.mach is None else "<br><br>"+record.mach.mach_name
            body += body_tmp.format(wo_no=record.wo_no, step_code=record.step_code, step_name=record.step_name,
                                    labor_time=record.labor_time, mach_time=record.mach_time, good_qty=record.good_qty,
                                    ng_qty=record.ng_qty, edit_btn=edit_btn, del_btn=del_btn,
                                    item_no=record.item_no, spec=record.spec, comment=comment, mach_name=mach_name)

        for record2 in record2s:
            if record2.qty == None:
                record2.qty = 0
            total_labor_time += record2.labor_time
            body_tmp = """<tr><td style='text-align:center' colspan='3'>{comment}</td><td>{type_code}</td><td>{type_name}</td>
                                        <td style='text-align:right'>{labor_time}</td>
                                        <td></td>
                                        <td style='text-align:right'>{qty}</td>
                                        <td></td>
                                        <td>{del_btn}</td></tr>"""
            if not request.user.is_anonymous:
                del_btn = """<a class="btn btn-danger m-1" href="/production/record2_delete/{record_pk}" role="button" onclick="return confirm('Are you sure?')">刪除</a>""".format(
                    record_pk=record2.pk)
            else:
                del_btn = ""
            body += body_tmp.format(type_code=record2.work_type.type_code, type_name=record2.work_type.type_name,
                                    labor_time=record2.labor_time, del_btn=del_btn, comment=record2.comment, qty=record2.qty)

        total_labor_time = round(total_labor_time, 1)
        rest_time = round(480 - total_labor_time, 1)
        if rest_time < 0:
            rest_time = 0

        option_tmp = ""

        for worktype in worktypes:
            type_name = worktype.type_name
            if lang == "vi":
                type_name = worktype.type_name_vi
            elif lang == "en":
                type_name = worktype.type_name_en

            option_tmp += """<option value="{value}">{name}</option>""".format(value=worktype.type_code,
                                                                               name=type_name)
        body += """<tr><td colspan='2'><select name="work_type" class="select custom-select" id="id_time_code" required>
                                <option value="">---------</option>
                                {options}
                            </select></td>""".format(options=option_tmp)
        body += """<td colspan='3'><input type="text" name="comment" placeholder="{comment}" class="textinput textInput form-control" id="id_comment"></td>""".format(comment=_('comment'))
        body += """<td><input type="text" name="labor_time" class="textinput textInput form-control" id="id_labor_time" required></td>"""
        body += """<td><input type="hidden" id="hid_record_dt2" name="hid_record_dt2" value='{record_dt2}'>
                <input type="hidden" id="hid_sap_emp_no" name="hid_sap_emp_no" value=""></td>""".format(record_dt2=record_dt)
        body += """<td><input type="text" name="qty" class="textinput textInput form-control" id="id_qty"></td><td></td>"""
        body += """<td><input type="submit" class="btn btn-success m-1" value="{new}" onclick="record2_submit()"></td></tr>""".format(new=_('New'))
        # 總人時計算
        body += """<tr><td colspan='5' style='text-align:right'>{trans_total_labor_time}</td><td style='text-align:right'>{total_labor_time}</td>
                        <td colspan='3' style='text-align:right'>{remain_labor_time}</td><td style='text-align:right'>{rest_time}</td></tr>""".format(
            total_labor_time=total_labor_time, rest_time=rest_time,
            trans_total_labor_time=_('total_labor_time'), remain_labor_time=_('remain_labor_time'))
        html = table.format(header=header, body=body)
        html += """<div style="text-align: right;color: #CCC">{username} {record_date}：{record_dt}</div>""".format(record_dt=record_dt, record_date=_('record_date'), username=username)
    form = RecordSearchForm(initial={'sap_emp_no': sap_emp_no, 'record_dt': record_dt})
    return render(request, 'production/record_detail.html', locals())


def get_step_info(request):
    value = {}
    if request.method == 'POST':
        cfm_code = request.POST.get('cfm_code')
        try:
            step = WODetail.objects.select_related('wo_main').get(wo_main__enable=True, cfm_code=cfm_code)
            labor_time = round(step.wo_labor_time/step.std_qty*step.wo_qty, 1)
            mach_time = round(step.wo_mach_time / step.std_qty * step.wo_qty, 1)

            if step:
                value['plant'] = step.wo_main.plant
                value['wo_no'] = step.wo_main.wo_no
                value['item_no'] = step.wo_main.item_no
                value['spec'] = step.wo_main.spec
                value['step_no'] = step.step_no
                value['step_code'] = step.step_code
                value['step_name'] = step.step_name
                value['labor_time'] = labor_time
                value['mach_time'] = mach_time
                value['good_qty'] = step.wo_qty
                value['wo_qty'] = step.wo_qty
                value['ctr_code'] = step.ctr_code
                value['work_center'] = step.work_center

                # 取得已報工數量及人時
                worked_good_qty = 0
                worked_ng_qty = 0
                worked_labor_time = 0
                worked_mach_time = 0
                worked_rows = Record.objects.filter(cfm_code=step.cfm_code).aggregate(Sum('labor_time'), Sum('mach_time'), Sum('good_qty'), Sum('ng_qty'))

                if worked_rows['good_qty__sum']:
                    worked_good_qty = worked_rows['good_qty__sum']
                if worked_rows['ng_qty__sum']:
                    worked_ng_qty = worked_rows['ng_qty__sum']
                if worked_rows['labor_time__sum']:
                    worked_labor_time = worked_rows['labor_time__sum']
                if worked_rows['mach_time__sum']:
                    worked_mach_time = worked_rows['mach_time__sum']

                value['worked_good_qty'] = worked_good_qty
                value['worked_ng_qty'] = worked_ng_qty
                value['worked_labor_time'] = worked_labor_time
                value['worked_mach_time'] = worked_mach_time

            # 判斷第一站是否已報工
            if step.step_no != "0010":
                records = Record.objects.filter(wo_no=value['wo_no'], step_no='0010')
                if records:
                    value['first_step_done'] = "Y"
                else:
                    value['first_step_done'] = "N"


            # 判斷是否輸入過用料
            consumptions = Consumption.objects.filter(cfm_code=cfm_code)
            if consumptions.exists():
                value['consumption_exist'] = "Y"
            else:
                value['consumption_exist'] = "N"
        except Exception as e:
            print(e)
    return JsonResponse(value, safe=False)


def get_user_info(request):
    value = {}
    if request.method == 'POST':
        try:
            sap_emp_no = request.POST.get('sap_emp_no')
            record_dt = request.POST.get('record_dt')
            key_user = CustomUser.objects.get(sap_emp_no=sap_emp_no)
            if key_user:
                username = key_user.username
                emp_no = key_user.emp_no
                value['sap_emp_no'] = sap_emp_no
                value['username'] = username
                value['emp_no'] = emp_no

            records = Record.objects.filter(record_dt=record_dt, emp_no=emp_no).aggregate(Sum('labor_time'), Sum('mach_time'))
            if records['labor_time__sum']:
                value['worked_labor_time'] = records['labor_time__sum']
            else:
                value['worked_labor_time'] = 0

        except Exception as e:
            print(e)

    return JsonResponse(value, safe=False)


def get_mach_info(request):
    value = {}
    html = ""
    if request.method == 'POST':
        try:
            step_code = request.POST.get('step_code')
            machs = Machine.objects.filter(step_code=step_code)
            if machs:
                html = "<option value='' selected=''>---------</option>"

            for mach in machs:
                html += "<option value='{mach_code}'>{mach_name}</option>".format(mach_code=mach.mach_code, mach_name=mach.mach_name)

            value["html"] = html
        except Exception as e:
            print(e)

    return JsonResponse(value, safe=False)


def get_mtr_info(request):
    value_list = []
    if request.method == 'POST':
        try:
            cfm_code = request.POST.get('cfm_code')
            records = Consumption.objects.filter(cfm_code=cfm_code)
            for record in records:
                value = {}
                value['wo_mtrl_no'] = record.wo_mtrl_no
                value['mtr_no'] = record.item_no
                value['qty'] = record.qty
                value_list.append(value)
            value = json.dumps(value_list)
        except Exception as e:
            print(e)

    return JsonResponse(value, safe=False)


def build_exceltemp_data(request, excel_file):
    # 刪除前先清空暫存表
    temp = ExcelTemp.objects.all()
    temp.delete()

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.worksheets[0]
    for iRow in range(2, sheet.max_row + 1):
        if not sheet.cell(row=iRow, column=1).value:
            break
        wo = {}
        wo['plant'] = sheet.cell(row=iRow, column=1).value
        wo['wo_no'] = sheet.cell(row=iRow, column=2).value
        wo['cfm_code'] = sheet.cell(row=iRow, column=3).value
        wo['step_no'] = sheet.cell(row=iRow, column=4).value
        wo['step_code'] = sheet.cell(row=iRow, column=5).value
        wo['step_name'] = sheet.cell(row=iRow, column=6).value
        wo['ctr_code'] = sheet.cell(row=iRow, column=7).value
        wo['status'] = sheet.cell(row=iRow, column=8).value
        wo['work_center'] = sheet.cell(row=iRow, column=9).value
        wo['item_no'] = sheet.cell(row=iRow, column=10).value
        wo['spec'] = sheet.cell(row=iRow, column=11).value
        wo['wo_qty'] = sheet.cell(row=iRow, column=12).value
        wo['wo_labor_time'] = sheet.cell(row=iRow, column=13).value
        wo['wo_mach_time'] = sheet.cell(row=iRow, column=14).value
        wo['std_qty'] = sheet.cell(row=iRow, column=15).value

        temp = ExcelTemp()
        temp.batch_no = uuid.uuid4().hex[:10]
        temp.plant = wo['plant']
        temp.wo_no = wo['wo_no']
        temp.work_center = wo['work_center']
        temp.item_no = wo['item_no']
        temp.spec = wo['spec']
        temp.cfm_code = wo['cfm_code']
        temp.step_no = wo['step_no']
        temp.step_code = wo['step_code']
        temp.step_name = wo['step_name']
        temp.ctr_code = wo['ctr_code']
        temp.status = wo['status']
        temp.wo_qty = wo['wo_qty']
        temp.wo_labor_time = wo['wo_labor_time']
        temp.wo_mach_time = wo['wo_mach_time']
        temp.std_qty = wo['std_qty']
        temp.create_by = request.user
        if str(wo['status']).find('PRT') >= 0:  # 已列印的工單才匯入系統
            temp.save()


@login_required
def excel_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        if excel_file:
            batch_no = uuid.uuid4().hex[:10]

            build_exceltemp_data(request, excel_file)

            rows = ExcelTemp.objects.values('plant', 'wo_no', 'item_no', 'spec').distinct()
            for row in rows:
                tmp = WOMain.objects.filter(wo_no=row['wo_no']).aggregate(version=Max('version'))
                if not tmp['version']:  # 新增資料
                    version = 1
                else:
                    WOMain.objects.filter(wo_no=row['wo_no'], enable=True).update(enable=False)
                    version = int(tmp['version']) + 1

                wo_main = WOMain()
                wo_main.id = uuid.uuid4().hex[:10]
                wo_main.batch_no = batch_no
                wo_main.item_no = row['item_no']
                wo_main.spec = row['spec']
                wo_main.plant = row['plant']
                wo_main.wo_no = row['wo_no']
                wo_main.version = version
                wo_main.create_by = request.user
                wo_main.save()

                tmp_details = ExcelTemp.objects.filter(wo_no=row['wo_no']).all()
                for tmp_detail in tmp_details:
                    detail = WODetail()
                    detail.wo_main = wo_main
                    detail.cfm_code = tmp_detail.cfm_code
                    detail.ctr_code = tmp_detail.ctr_code
                    detail.status = tmp_detail.status
                    detail.work_center = tmp_detail.work_center
                    detail.step_no = tmp_detail.step_no
                    detail.step_code = tmp_detail.step_code
                    detail.step_name = tmp_detail.step_name
                    detail.wo_qty = tmp_detail.wo_qty
                    detail.wo_labor_time = tmp_detail.wo_labor_time
                    detail.wo_mach_time = tmp_detail.wo_mach_time
                    detail.std_qty = tmp_detail.std_qty
                    detail.save()
            save_path = os.path.join(settings.MEDIA_ROOT, 'uploads', 'production', 'coois', excel_file.name)
            file_name = default_storage.save(save_path, excel_file)

            coois = COOIS_Record()
            coois.batch_no = batch_no
            coois.file_name = file_name
            coois.file_url = "/media/uploads/production/coois/"+file_name[file_name.rfind('/')+1:]
            coois.create_by = request.user
            coois.save()

    cooises = COOIS_Record.objects.all().order_by('-create_at')[:20]
    for coois in cooises:
        coois.file_name = coois.file_name[coois.file_name.rfind('/')+1:]


    return render(request, 'production/import.html', locals())


#API預覽
def excel_import_preview(request):
    wos = []
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        if excel_file:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.worksheets[0]
            for iRow in range(2, sheet.max_row+1):
                wo = {}
                wo['plant'] = sheet.cell(row=iRow, column=1).value
                wo['wo_no'] = sheet.cell(row=iRow, column=2).value
                wo['cfm_code'] = sheet.cell(row=iRow, column=3).value
                wo['step_no'] = sheet.cell(row=iRow, column=4).value
                wo['step_code'] = sheet.cell(row=iRow, column=5).value
                wo['step_name'] = sheet.cell(row=iRow, column=6).value
                wo['ctr_code'] = sheet.cell(row=iRow, column=7).value
                wo['status'] = sheet.cell(row=iRow, column=8).value
                wo['work_center'] = sheet.cell(row=iRow, column=9).value
                wo['item_no'] = sheet.cell(row=iRow, column=10).value
                wo['spec'] = sheet.cell(row=iRow, column=11).value
                wo['wo_qty'] = sheet.cell(row=iRow, column=12).value
                wo['wo_labor_time'] = sheet.cell(row=iRow, column=13).value
                wo['wo_mach_time'] = sheet.cell(row=iRow, column=14).value
                wo['std_qty'] = sheet.cell(row=iRow, column=15).value
                wos.append(wo)

                if iRow > 50:  # 預覽只顯示50筆
                    break

            html = COOIS2Table(wos)
    return JsonResponse(html, safe=False)


def record_manage(request):
    form = RecordManageForm()
    if request.method == 'POST':
        record_dt = request.POST.get('record_dt')
        form = RecordManageForm(initial={'record_dt': record_dt})
    else:
        now = datetime.now()
        record_dt = datetime.strftime(now, '%Y-%m-%d')

    if request.user.is_superuser or request.user.unit.id == 24:
        record1_rows = Record.objects.filter(record_dt=record_dt).values('sap_emp_no').distinct()
        record2_rows = Record2.objects.filter(record_dt=record_dt).values('sap_emp_no').distinct()
    else:
        # 取得同部門人員清單
        dept_users = CustomUser.objects.filter(unit=request.user.unit)
        dept_users = [dept_user.sap_emp_no for dept_user in dept_users]
        record1_rows = Record.objects.filter(record_dt=record_dt, sap_emp_no__in=dept_users).values('sap_emp_no').distinct()
        record2_rows = Record2.objects.filter(record_dt=record_dt, sap_emp_no__in=dept_users).values('sap_emp_no').distinct()

    list = []  # 當日報工人員列表
    for record1_row in record1_rows:
        list.append(record1_row['sap_emp_no'])

    for record2_row in record2_rows:
        if not record2_row['sap_emp_no'] in list:
            list.append(record2_row['sap_emp_no'])

    # 人員報工統計
    records = []
    for sap_emp_no in list:
        record = {}
        css_color = ""
        html = ""
        count = 0
        result = Record.objects.filter(record_dt=record_dt, sap_emp_no=sap_emp_no).aggregate(Sum('labor_time'), Sum('mach_time'))
        record['labor_time'] = 0
        if result['labor_time__sum']:
            record['labor_time'] = round(result['labor_time__sum'], 1)
        record['mach_time'] = 0
        if result['mach_time__sum']:
            record['mach_time'] = round(result['mach_time__sum'], 1)

        # 人時加總
        count += record['labor_time']

        key_user = CustomUser.objects.get(sap_emp_no=sap_emp_no)
        record['sap_emp_no'] = sap_emp_no
        record['username'] = key_user.username
        worktypes = WorkType.objects.all().order_by('type_code')
        for worktype in worktypes:
            result = Record2.objects.filter(record_dt=record_dt, sap_emp_no=sap_emp_no, work_type=worktype).aggregate(Sum('labor_time'))
            if not result['labor_time__sum']:
                tmp = 0
            else:
                tmp = float(result['labor_time__sum'])
            count += tmp  # 人時加總
            html += """<td>{value}</td>""".format(value=tmp)
        record['html'] = html
        record['count'] = count
        records.append(record)

        for record in records:
            for worktype in worktypes:
                html += """<td>{% record[worktype.type_name] %} </td>"""

        if count > 480:
            record['css_overtime'] = "color: red;"
    return render(request, 'production/record_manage.html', locals())


def prod_sap_export(request):
    dc_db = dc_database()
    sqlite_db = database()
    save_path = MEDIA_ROOT + 'sync_sap_excel\\'
    work_sync = SYN_Noah_WorkHour(sqlite_db, dc_db, request.user, save_path)
    mtrl_sync = SYN_Noah_Consumption(sqlite_db, dc_db, request.user, save_path)

    if request.method == 'POST':
        plant = request.POST.get('plant')
        action = request.POST.get('action')

        if action == "workhour":  # 工時
            sync = work_sync

        if action == "consumption":  # 物料
            sync = mtrl_sync

        file_name = sync.get_file_name(plant)
        file_path = save_path + file_name
        amount = sync.generate_csv(plant, file_path, file_name)

        if amount > 0:
            data = open(file_path, 'r').read()
            response = HttpResponse(data, content_type='text/csv')
            response['Content-Disposition'] = "attachment; filename={file_name}".format(file_name=file_name)
            return response

    logs = Sync_SAP_Log.objects.all().order_by('-create_at')[:20]

    record_list = []
    for plant in ['302A', '302B']:
        record_count = {}
        work_records = work_sync.export_records(plant)
        mtrl_records = mtrl_sync.export_consumptions(plant)
        record_count['plant'] = plant
        record_count['record_count'] = len(work_records)
        record_count['material_count'] = len(mtrl_records)
        record_list.append(record_count)

    form = ExportForm()
    return render(request, 'production/export.html', locals())

#Excel
def record_export(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        due_date = request.POST.get('due_date')

        db = database()
        sql = """select plant,wo_no,user.emp_no,unit.unitName,user.username,r.step_code,r.step_name,record_dt,labor_time,m.mach_name,mach_time,good_qty,ng_qty,comment,r.update_at,r.work_center 
                    from production_record r,users_customuser user, users_unit unit left outer join production_machine m on  r.mach_id = m.mach_code 
                    where r.sap_emp_no = user.sap_emp_no and user.unit_id = unit.id 
                    and record_dt between '{start_date}' and '{due_date}'
                    union
                    select '','',user.emp_no,unit.unitName,user.username,w.type_code,w.type_name,record_dt,labor_time,'','',qty,'',comment, r.create_at,'' 
                    from production_record2 r,users_customuser user, users_unit unit , production_worktype w
                    where r.sap_emp_no=user.sap_emp_no and user.unit_id = unit.id and r.work_type_id = w.type_code
                    and record_dt between '{start_date}' and '{due_date}'
                    union
                    select plant,'',emp_no,unitName,username,'','橡膠成型機時計算' step_name,record_dt,max(mach_time) mach_time,'','','','','','','' from (
                    select r.plant,r.record_dt,r.emp_no,unit.unitName,user.username,mach_name,sum(mach_time) mach_time from production_record r,users_customuser user, users_unit unit, production_machine m  
                    where r.sap_emp_no = user.sap_emp_no and user.unit_id = unit.id and r.mach_id = m.mach_code and r.step_code = 'TWA027'
                    and record_dt between '{start_date}' and '{due_date}'
                    group by r.plant,r.record_dt,r.emp_no,unit.unitName,user.username,mach_name
                    ) A group by plant,record_dt,emp_no,username
                    order by unitName, username, record_dt""".format(start_date=start_date, due_date=due_date)
        records = db.select_sql_dict(sql)

        file_name = "Record_{record_dt}.xls".format(record_dt=start_date)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = "attachment; filename={file_name}".format(file_name=file_name)

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('報工紀錄')
        ws.col(0).width = 256 * 20
        ws.col(1).width = 256 * 20
        ws.col(2).width = 256 * 20
        ws.col(3).width = 256 * 20
        ws.col(4).width = 256 * 20
        ws.col(5).width = 256 * 20
        ws.col(6).width = 256 * 20
        ws.col(7).width = 256 * 20
        ws.col(8).width = 256 * 20
        ws.col(9).width = 256 * 20
        ws.col(10).width = 256 * 20
        ws.col(11).width = 256 * 20
        ws.col(12).width = 256 * 20
        ws.col(13).width = 256 * 20
        ws.col(14).width = 256 * 20
        ws.col(15).width = 256 * 20

        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['廠別', '工單', '部門', 'EMP NO', '姓名', '站點碼',
                   '站點名稱', '報工日期', '人時', '機台', '機時', '良品數量', 'NG數量', 'Comment', '紀錄時間', 'Work Center']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'yyyy/mm/dd'

        for record in records:
            row_num += 1

            ws.write(row_num, 0, record['plant'], font_style)  # 廠別
            ws.write(row_num, 1, record['wo_no'], font_style)  # 工單
            ws.write(row_num, 2, record['unitName'], font_style)  # 部門
            ws.write(row_num, 3, record['emp_no'], font_style)  # EMP NO
            ws.write(row_num, 4, record['username'], font_style)  # 姓名
            ws.write(row_num, 5, record['step_code'], font_style)  # 站點碼
            ws.write(row_num, 6, record['step_name'], font_style)  # 站點名稱
            ws.write(row_num, 7, record['record_dt'], font_style)  # 報工日期
            ws.write(row_num, 8, record['labor_time'], font_style)  # 人時
            ws.write(row_num, 9, record['mach_name'], font_style)  # 機台
            ws.write(row_num, 10, record['mach_time'], font_style)  # 機時
            ws.write(row_num, 11, record['good_qty'], font_style)  # 良品數量
            ws.write(row_num, 12, record['ng_qty'], font_style)  # NG數量
            ws.write(row_num, 13, record['comment'], font_style)  # Comment
            ws.write(row_num, 14, record['update_at'], date_format)  # 紀錄時間
            ws.write(row_num, 15, record['work_center'], font_style)  # Work Center

        wb.save(response)
        return response
    form = RecordHistoryForm()
    return render(request, 'production/record_excel.html', locals())



